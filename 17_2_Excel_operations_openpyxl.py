from openpyxl import Workbook, load_workbook
import os

excel_file = "students.xlsx"
# Create a new Excel file
if not os.path.exists(excel_file):
    workbook = Workbook()

    # Create a worksheet

    ws = workbook.active


    # print datas

    ws["A1"] = "Name"
    ws["B1"] = "Surname"
    ws["C1"] = "Age"
else:
    workbook = load_workbook(excel_file)
    ws = workbook.active


# Add new rows

ws.append(["Yakubi", "Acar", 16])
ws.append(["Muhammet Taha", "Acar", 14])

workbook.save("./students.xlsx")
