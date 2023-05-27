from module.browser import Browser
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import logging
import sys
import os
import urllib.request
from openpyxl import Workbook, load_workbook
from datetime import datetime

# If there is no folder to download the pictures, create it
images_dir = "./images"
if not os.path.exists(images_dir):
    os.makedirs(images_dir)

# Create excel file if not exists

excel_path = "amazon_products.xlsx"
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    # print title
    ws.append(["Date", "ID", "NAME", "IMAGE", "PRİCE"])

else:
    wb = load_workbook(excel_path)
    ws = wb.active

cmd_log = logging.StreamHandler(stream=sys.stdout)

logging.basicConfig(
    handlers=[cmd_log],
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s ",
    level=logging.INFO,
    datefmt="%Y.%m.%d - %H:%M:%S"
)

logger = logging.getLogger(__name__)

driver = Browser().get()

# Move to amazon page

driver.get("https://www.amazon.com.tr/")
wait = WebDriverWait(driver, 30)
while True:
    try:
        wait.until(EC.visibility_of_element_located((By.ID, "twotabsearchtextbox")))
        break
    except:
        driver.refresh()

# if cookies exist close that pop-up
try:
    driver.find_element(By.ID, "sp-cc-accept").click()
except:
    logger.error("Pop-up could not find")

# Search
driver.minimize_window()
search = "Laptop Sırt Çantası"
# input("What you want to search: ")
driver.maximize_window()
driver.find_element(By.ID, "twotabsearchtextbox").send_keys(search)
driver.find_element(By.ID, "nav-search-submit-button").click()

# browse products

products = driver.find_elements(By.XPATH, "//div[@data-component-type='s-search-result']")
for product in products:
    # get the id of the product
    id = product.get_attribute("data-asin")
    logger.info(f"{id} id'li ürün bilgileri alınıyor")

    # get the name of the product
    name = product.find_element(By.TAG_NAME, "h2").text
    logger.info(f"name of product is {name.upper()}")

    # get the image link of the product
    src = product.find_element(By.TAG_NAME, "img").get_attribute("src")
    image_path = f"{images_dir}/{id}.jpg"
    urllib.request.urlretrieve(src, f"{images_dir}/{id}.jpg")
    logger.info(f"link of image is {src} (downloaded)")

    # get the price of the product
    try:
        lira = product.find_element(By.CLASS_NAME, "a-price-whole").text
        kurus = product.find_element(By.CLASS_NAME, "a-price-fraction").text
        price = float(f"{lira}.{kurus}")
        logger.info(f"Price of the product is {price} TL")
    except:
        price = ""
        logger.error("Price of product could not find")

    # Excel entry procedures
    ws.append([
        datetime.now(),
        id,
        name,
        f'=HYPERLINK("{os.getcwd()}/{image_path}", "IMAGE")',
        price
    ])

wb.save(excel_path)
